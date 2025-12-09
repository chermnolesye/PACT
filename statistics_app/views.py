from django.shortcuts import render
from core_app.models import (
    Text,
    Token,
    Error,
    Error,
	ErrorLevel,
    ErrorToken,
    ErrorTag,
    Group,
	Emotion,
    AcademicYear,
    TextType,
    Student,
    Sentence,
    User,
)

from collections import defaultdict
from statistics_app import dashboards
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q, F, Value, IntegerField
import json
import openpyxl
from openpyxl.styles import Font
import numpy as np
import scipy

def export_group_error_stats(request):
    group_id = request.GET.get('group')
    if not group_id:
        return HttpResponse("Группа не выбрана", status=400)

    try:
        group = Group.objects.get(idgroup=group_id)
    except Group.DoesNotExist:
        return HttpResponse("Группа не найдена", status=404)

    students = Student.objects.filter(idgroup=group_id).select_related("iduser").order_by("iduser__lastname")
    tags = {t.iderrortag: t.tagtext for t in ErrorTag.objects.all()}

    wb = openpyxl.Workbook()
    default_sheet = wb.active 

    added_sheets = False  

    for student in students:
        text_ids = Text.objects.filter(idstudent=student.idstudent).values_list('idtext', flat=True)

        error_counts_raw = (
            Error.objects
            .filter(errortoken__idtoken__idsentence__idtext__in=text_ids)
            .values("iderrortag")
            .annotate(count=Count("iderror"))
        )

        if not error_counts_raw:
            continue

        error_counts = []
        for error in error_counts_raw:
            tag_id = error["iderrortag"]
            tag_name = tags.get(tag_id, "Неизвестно")
            count = error["count"]
            error_counts.append((tag_name, count))
        error_counts.sort(key=lambda x: x[0])

        sheet_name = f"{student.iduser.lastname} {student.iduser.firstname} {student.iduser.middlename}"
        sheet_name = sheet_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        ws.append(["Тэг", "Частота"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for tag_name, count in error_counts:
            ws.append([tag_name, count])

        added_sheets = True

    if added_sheets:
        wb.remove(default_sheet)
    else:
        default_sheet.title = "Нет данных"
        default_sheet.append(["Нет ошибок у студентов в этой группе."])

    filename = f"{group.groupname} ({group.idayear.title}).xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def statistics_view(request):
    groups = (
        Group.objects.select_related("idayear")
        .all()
        .values("idgroup", "groupname", "idayear__title")
        .distinct()
    )
    group_data = [
        {
            "id": group["idgroup"],
            "name": group["groupname"],
            "year": group["idayear__title"],
        }
        for group in groups
    ]

    group_id = ""

    context = {"groups": group_data, "selected_group": group_id}

    return render(request, "statistics.html", context)


def error_stats(request):
    tags = ErrorTag.objects.all().values(
        "iderrortag",
        "tagtext",
        "tagtextrussian",
        "idtagparent",
    )

    error_counts = Error.objects.values("iderrortag", "iderrorlevel").annotate(
        count=Count("iderror")
    )

    error_count_map = defaultdict(lambda: defaultdict(int))
    for item in error_counts:
        tag_id = item["iderrortag"]
        level_id = item["iderrorlevel"]
        count = item["count"]
        error_count_map[tag_id][level_id] = count

    grouped_tags = defaultdict(list)

    for tag in tags:
        tag_id = tag["iderrortag"]
        tag_name = tag["tagtext"]
        parent_id = tag["idtagparent"]

        levels = error_count_map.get(tag_id, {})

        active_levels = [lvl for lvl, cnt in levels.items() if cnt > 0]
        num_active_levels = len(active_levels)

        if num_active_levels == 0 or num_active_levels == 1:
            color = "#ffffff"
        elif num_active_levels == 2:
            color = "#e8f0fe"
        else:
            color = "#cfe2ff"

        tag_info = {
            "id": tag_id,
            "nametag": tag_name,
            "color": color,
            "level1": levels.get(1, 0),
            "level2": levels.get(2, 0),
            "level3": levels.get(3, 0),
            "parent_id": parent_id, 
        }

        if parent_id:
            parent_name = next(
                (t["tagtext"] for t in tags if t["iderrortag"] == parent_id), None
            )
            if parent_name:
                grouped_tags[parent_name].append(tag_info)
        else:
            grouped_tags[tag_name].append(tag_info)  

    context = {"tags_error": dict(grouped_tags)}
    return render(request, "error_stats.html", context)


def chart_types_errors(request):
    if request.method != "POST":
        levels = dashboards.get_levels()

        groups = list(
            Group.objects.values("idgroup", "groupname", "idayear__title")
            .distinct()
            .order_by("groupname")
        )

        courses = list(
            Group.objects.values("studycourse")
            .filter(studycourse__gt=0)
            .distinct()
            .order_by("studycourse")
        )

        texts = list(
            Text.objects.values("header")
            .filter(errorcheckflag=True)
            .distinct()
            .order_by("header")
        )

        text_types = list(
            TextType.objects.filter(text__errorcheckflag=True)
            .distinct()
            .order_by("idtexttype")
            .values()
        )

        data_count_errors = list(
            ErrorToken.objects.values(
                "iderror__iderrortag__iderrortag",
                "iderror__iderrortag__idtagparent",
                "iderror__iderrortag__tagtext",
                "iderror__iderrortag__tagtextrussian",
                "iderror__iderrortag__tagcolor",
                "idtoken__idsentence__idtext",
            )
            .filter(
                Q(iderror__iderrortag__isnull=False) &
                Q(idtoken__idsentence__idtext__errorcheckflag=True)
            )
            .annotate(count_data=Count("iderror__iderrortag__iderrortag"))
        )

        data_on_tokens = dashboards.get_data_on_tokens(
            data_count_errors,
            id_data="iderror__iderrortag__iderrortag",
            is_unique_data=True,
            is_for_one_group=False,
        )

        data = dashboards.get_data_errors(data_on_tokens, level=0, is_sorted=True)

        tag_parents, dict_children = dashboards.get_dict_children()

        return render(
            request,
            "dashboard_error_types.html",
            {
                "right": True,
                "levels": levels,
                "groups": groups,
                "courses": courses,
                "texts": texts,
                "text_types": text_types,
                "data": data,
                "tag_parents": tag_parents,
                "dict_children": dict_children,
            },
        )

    else:
        list_filters = json.loads(request.body)
        flag_post = list_filters["flag_post"]

        if flag_post == "enrollment_date":
            enrollment_date = dashboards.get_enrollment_date(list_filters)
            return JsonResponse({"enrollment_date": enrollment_date}, status=200)

        if flag_post == "choice_all":
            texts, text_types = dashboards.get_filters_for_choice_all(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_group":
            group_name = list_filters.get("group")
            enrollment_date = list_filters.get("enrollment_date")
            
            academic_year_title = None
            if enrollment_date:
                if " \\ " in enrollment_date:
                    start_year = enrollment_date.split(" \\ ")[0]
                    academic_year_title = f"{start_year}/{int(start_year) + 1}"
                else:
                    academic_year_title = enrollment_date
            
            texts = list(
                Text.objects.filter(
                    errorcheckflag=True,
                    idstudent__idgroup__groupname=group_name,
                    idstudent__idgroup__idayear__title=academic_year_title
                )
                .values("header")
                .distinct()
                .order_by("header")
            )
            
            text_types = list(
                TextType.objects.filter(
                    text__errorcheckflag=True,
                    text__idstudent__idgroup__groupname=group_name,
                    text__idstudent__idgroup__idayear__title=academic_year_title
                )
                .distinct()
                .order_by("idtexttype")
                .values()
            )
            
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_student":
            surname = list_filters.get("surname")
            name = list_filters.get("name")
            patronymic = list_filters.get("patronymic")
            
            base_filter = Text.objects.filter(
                errorcheckflag=True,
                idstudent__iduser__lastname__iexact=surname,
                idstudent__iduser__firstname__iexact=name
            )
            
            if patronymic:
                base_filter = base_filter.filter(
                    idstudent__iduser__middlename__iexact=patronymic
                )
            
            texts = list(
                base_filter.values("header")
                .distinct()
                .order_by("header")
            )
            
            text_types = list(
                TextType.objects.filter(
                    text__errorcheckflag=True,
                    text__idstudent__iduser__lastname__iexact=surname,
                    text__idstudent__iduser__firstname__iexact=name
                )
                .distinct()
                .order_by("idtexttype")
                .values()
            )
            
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_course":
            course = list_filters.get("course")
            
            texts = list(
                Text.objects.filter(
                    errorcheckflag=True,
                    idstudent__idgroup__studycourse=course
                )
                .values("header")
                .distinct()
                .order_by("header")
            )
            
            text_types = list(
                TextType.objects.filter(
                    text__errorcheckflag=True,
                    text__idstudent__idgroup__studycourse=course
                )
                .distinct()
                .order_by("idtexttype")
                .values()
            )
            
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_text":
            groups, courses, text_types = dashboards.get_filters_for_choice_text(
                list_filters
            )
            return JsonResponse(
                {"groups": groups, "courses": courses, "text_types": text_types},
                status=200,
            )

        if flag_post == "choice_text_type":
            groups, courses, texts = dashboards.get_filters_for_choice_text_type(
                list_filters
            )
            return JsonResponse(
                {"groups": groups, "courses": courses, "texts": texts}, status=200
            )

        if flag_post == "update_diagrams":
            group = list_filters.get("group")
            date = list_filters.get("enrollment_date")
            surname = list_filters.get("surname")
            name = list_filters.get("name")
            patronymic = list_filters.get("patronymic")
            course = list_filters.get("course")
            text = list_filters.get("text")
            text_type = list_filters.get("text_type")
            level = int(list_filters.get("level", 0))

            academic_year_title = None
            if date:
                if " \\ " in date:
                    start_year = date.split(" \\ ")[0]
                    academic_year_title = f"{start_year}/{int(start_year) + 1}"
                else:
                    academic_year_title = date

            base_filter = Q(iderror__iderrortag__isnull=False) & Q(
                idtoken__idsentence__idtext__errorcheckflag=True
            )

            if surname and name:
                base_filter &= Q(
                    idtoken__idsentence__idtext__idstudent__iduser__lastname__iexact=surname
                ) & Q(
                    idtoken__idsentence__idtext__idstudent__iduser__firstname__iexact=name
                )
                if patronymic:
                    base_filter &= Q(
                        idtoken__idsentence__idtext__idstudent__iduser__middlename__iexact=patronymic
                    )

            if course:
                base_filter &= Q(
                    idtoken__idsentence__idtext__idstudent__idgroup__studycourse=course
                )

            if group and academic_year_title:
                base_filter &= Q(
                    idtoken__idsentence__idtext__idstudent__idgroup__groupname=group
                ) & Q(
                    idtoken__idsentence__idtext__idstudent__idgroup__idayear__title=academic_year_title
                )

            if text:
                base_filter &= Q(idtoken__idsentence__idtext__header=text)

            if text_type:
                base_filter &= Q(idtoken__idsentence__idtext__idtexttype__texttypename=text_type)

            data_count_errors = list(
                ErrorToken.objects.filter(base_filter)
                .values(
                    "iderror__iderrortag__iderrortag",
                    "iderror__iderrortag__idtagparent",
                    "iderror__iderrortag__tagtext",
                    "iderror__iderrortag__tagtextrussian",
                    "idtoken__idsentence__idtext",
                )
                .annotate(count_data=Count("iderror__iderrortag__iderrortag"))
            )

            data_on_tokens = dashboards.get_data_on_tokens(
                data_count_errors, "iderror__iderrortag__iderrortag", True, False
            )
            data = dashboards.get_data_errors(data_on_tokens, level, True)

            return JsonResponse({"data_type_errors": data}, status=200)

def chart_grade_errors(request):
    if request.method != 'POST':
        languages = ['Deustache']
        groups = list(Group.objects.values('idgroup', 'groupname', 'idayear__title')
                     .distinct().order_by('groupname', 'idayear__title'))
        courses = list(
            Group.objects.values('studycourse').filter(studycourse__gt=0).distinct().order_by('studycourse'))
        texts = list(
            Text.objects.values('header').filter(errorcheckflag=True).distinct().order_by('header'))
        text_types = list(
            TextType.objects.values('idtexttype','texttypename').filter(text__errorcheckflag=True).distinct().order_by('idtexttype'))

        data_errorlevel = list(Error.objects.values(
            'iderrorlevel__iderrorlevel', 
            'iderrorlevel__errorlevelname',
            'iderrorlevel__errorlevelrussian'
        ).filter(
            errortoken__idtoken__idsentence__idtext__errorcheckflag=True,
            iderrorlevel__isnull=False
        ).annotate(count_data=Count('iderror', distinct=True)))

        total_tokens = Token.objects.filter(
            idsentence__idtext__errorcheckflag=True
        ).count()
        
        for item in data_errorlevel:
            if total_tokens > 0:
                item['count_data_on_tokens'] = (item['count_data'] * 100 / total_tokens)
            else:
                item['count_data_on_tokens'] = 0

        all_levels = ErrorLevel.objects.all()
        existing_level_ids = [item['iderrorlevel__iderrorlevel'] for item in data_errorlevel]
        
        for level in all_levels:
            if level.iderrorlevel not in existing_level_ids:
                data_errorlevel.append({
                    'iderrorlevel__iderrorlevel': level.iderrorlevel,
                    'iderrorlevel__errorlevelname': level.errorlevelname,
                    'iderrorlevel__errorlevelrussian': level.errorlevelrussian, 
                    'count_data': 0,
                    'count_data_on_tokens': 0
                })

        data_errorlevel = sorted(data_errorlevel, key=lambda d: d['count_data'], reverse=True)

        return render(request, 'dashboard_error_grade.html', {
            'right': True, 
            'languages': languages, 
            'groups': groups,
            'courses': courses, 
            'texts': texts,
            'text_types': text_types, 
            'data': data_errorlevel
        })
    else:
        list_filters = json.loads(request.body)
        flag_post = list_filters['flag_post']
        
        if flag_post == 'enrollment_date':
            enrollment_date = dashboards.get_enrollment_date(list_filters)
            return JsonResponse({'enrollment_date': enrollment_date}, status=200)
            
        if flag_post == 'choice_all':
            texts = list(Text.objects.values('header').filter(errorcheckflag=True).distinct().order_by('header'))
            text_types = list(TextType.objects.values('idtexttype','texttypename').filter(text__errorcheckflag=True).distinct().order_by('idtexttype'))
            return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
            
        if flag_post == 'choice_group':
            group = list_filters.get('group')
            date = list_filters.get('enrollment_date')
            
            texts_qs = Text.objects.filter(errorcheckflag=True)
            if group and date:
                texts_qs = texts_qs.filter(
                    idstudent__idgroup__groupname=group,
                    idstudent__idgroup__idayear__title=date
                )
            elif group:
                texts_qs = texts_qs.filter(
                    idstudent__idgroup__groupname=group
                )
            
            texts = list(texts_qs.values('header').distinct().order_by('header'))
            text_types = list(TextType.objects.values('idtexttype','texttypename')
                             .filter(text__in=texts_qs).distinct().order_by('idtexttype'))
            
            return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
            
        if flag_post == 'choice_student':
            surname = list_filters.get('surname')
            name = list_filters.get('name')
            patronymic = list_filters.get('patronymic')
            
            texts_qs = Text.objects.filter(errorcheckflag=True)
            if surname and name and patronymic:
                texts_qs = texts_qs.filter(
                    idstudent__iduser__lastname=surname,
                    idstudent__iduser__firstname=name,
                    idstudent__iduser__middlename=patronymic
                )
            elif surname and name:
                texts_qs = texts_qs.filter(
                    idstudent__iduser__lastname=surname,
                    idstudent__iduser__firstname=name
                )
            
            texts = list(texts_qs.values('header').distinct().order_by('header'))
            text_types = list(TextType.objects.values('idtexttype','texttypename')
                             .filter(text__in=texts_qs).distinct().order_by('idtexttype'))
            
            return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
            
        if flag_post == 'choice_course':
            course = list_filters.get('course')
            
            texts_qs = Text.objects.filter(errorcheckflag=True)
            if course:
                texts_qs = texts_qs.filter(
                    idstudent__idgroup__studycourse=course
                )
            
            texts = list(texts_qs.values('header').distinct().order_by('header'))
            text_types = list(TextType.objects.values('idtexttype','texttypename')
                             .filter(text__in=texts_qs).distinct().order_by('idtexttype'))
            
            return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
            
        if flag_post == 'choice_text':
            text = list_filters.get('text')
            
            texts_qs = Text.objects.filter(header=text, errorcheckflag=True)
            
            groups = list(Group.objects.values('idgroup', 'groupname', 'idayear__title')
                         .filter(student__text__in=texts_qs).distinct().order_by('groupname', 'idayear__title'))
            courses = list(Group.objects.values('studycourse')
                          .filter(student__text__in=texts_qs, studycourse__gt=0).distinct().order_by('studycourse'))
            text_types = list(TextType.objects.values('idtexttype','texttypename')
                             .filter(text__in=texts_qs).distinct().order_by('idtexttype'))
            
            return JsonResponse({'groups': groups, 'courses': courses, 'text_types': text_types}, status=200)
            
        if flag_post == 'choice_text_type':
            text_type = list_filters.get('text_type')
            
            texts_qs = Text.objects.filter(idtexttype__texttypename=text_type, errorcheckflag=True)
            
            groups = list(Group.objects.values('idgroup', 'groupname', 'idayear__title')
                         .filter(student__text__in=texts_qs).distinct().order_by('groupname', 'idayear__title'))
            courses = list(Group.objects.values('studycourse')
                          .filter(student__text__in=texts_qs, studycourse__gt=0).distinct().order_by('studycourse'))
            texts = list(texts_qs.values('header').distinct().order_by('header'))
            
            return JsonResponse({'groups': groups, 'courses': courses, 'texts': texts}, status=200)
            
        
        if flag_post == 'update_diagrams':
            group = list_filters.get('group')
            date = list_filters.get('enrollment_date')
            surname = list_filters.get('surname')
            name = list_filters.get('name')
            patronymic = list_filters.get('patronymic')
            course = list_filters.get('course')
            text = list_filters.get('text')
            text_type = list_filters.get('text_type')
            
            errors_qs = Error.objects.filter(
                errortoken__idtoken__idsentence__idtext__errorcheckflag=True,
                iderrorlevel__isnull=False
            )
            
            if surname and name and patronymic:
                errors_qs = errors_qs.filter(
                    errortoken__idtoken__idsentence__idtext__idstudent__iduser__lastname=surname,
                    errortoken__idtoken__idsentence__idtext__idstudent__iduser__firstname=name,
                    errortoken__idtoken__idsentence__idtext__idstudent__iduser__middlename=patronymic
                )
            elif surname and name:
                errors_qs = errors_qs.filter(
                    errortoken__idtoken__idsentence__idtext__idstudent__iduser__lastname=surname,
                    errortoken__idtoken__idsentence__idtext__idstudent__iduser__firstname=name
                )
            
            if group and date:
                errors_qs = errors_qs.filter(
                    errortoken__idtoken__idsentence__idtext__idstudent__idgroup__groupname=group,
                    errortoken__idtoken__idsentence__idtext__idstudent__idgroup__idayear__title=date
                )
            elif group:
                errors_qs = errors_qs.filter(
                    errortoken__idtoken__idsentence__idtext__idstudent__idgroup__groupname=group
                )
            
            if course:
                errors_qs = errors_qs.filter(
                    errortoken__idtoken__idsentence__idtext__idstudent__idgroup__studycourse=course
                )
            
            if text:
                errors_qs = errors_qs.filter(
                    errortoken__idtoken__idsentence__idtext__header=text
                )
            
            if text_type:
                errors_qs = errors_qs.filter(
                    errortoken__idtoken__idsentence__idtext__idtexttype__texttypename=text_type
                )
            
            data_errorlevel = list(
                errors_qs.values(
                    'iderrorlevel__iderrorlevel', 
                    'iderrorlevel__errorlevelname',
                    'iderrorlevel__errorlevelrussian'  
                ).annotate(count_data=Count('iderror', distinct=True))
            )
            
            tokens_qs = Token.objects.filter(
                idsentence__idtext__errorcheckflag=True
            )
            
            if surname and name and patronymic:
                tokens_qs = tokens_qs.filter(
                    idsentence__idtext__idstudent__iduser__lastname=surname,
                    idsentence__idtext__idstudent__iduser__firstname=name,
                    idsentence__idtext__idstudent__iduser__middlename=patronymic
                )
            elif surname and name:
                tokens_qs = tokens_qs.filter(
                    idsentence__idtext__idstudent__iduser__lastname=surname,
                    idsentence__idtext__idstudent__iduser__firstname=name
                )
            
            if group and date:
                tokens_qs = tokens_qs.filter(
                    idsentence__idtext__idstudent__idgroup__groupname=group,
                    idsentence__idtext__idstudent__idgroup__idayear__title=date
                )
            elif group:
                tokens_qs = tokens_qs.filter(
                    idsentence__idtext__idstudent__idgroup__groupname=group
                )
            
            if course:
                tokens_qs = tokens_qs.filter(
                    idsentence__idtext__idstudent__idgroup__studycourse=course
                )
            
            if text:
                tokens_qs = tokens_qs.filter(
                    idsentence__idtext__header=text
                )
            
            if text_type:
                tokens_qs = tokens_qs.filter(
                    idsentence__idtext__idtexttype__texttypename=text_type
                )
            
            total_tokens = tokens_qs.count()
            
            for item in data_errorlevel:
                if total_tokens > 0:
                    item['count_data_on_tokens'] = (item['count_data'] * 100 / total_tokens)
                else:
                    item['count_data_on_tokens'] = 0

            all_levels = ErrorLevel.objects.all()
            existing_level_ids = [item['iderrorlevel__iderrorlevel'] for item in data_errorlevel]
            
            for level in all_levels:
                if level.iderrorlevel not in existing_level_ids:
                    data_errorlevel.append({
                        'iderrorlevel__iderrorlevel': level.iderrorlevel,
                        'iderrorlevel__errorlevelname': level.errorlevelname,
                        'iderrorlevel__errorlevelrussian': level.errorlevelrussian,  
                        'count_data': 0,
                        'count_data_on_tokens': 0
                    })

            data_errorlevel = sorted(data_errorlevel, key=lambda d: d['count_data'], reverse=True)
            
            return JsonResponse({'data_grade_errors': data_errorlevel}, status=200)

def chart_types_grade_errors(request):		
    if request.method != 'POST':
        print("=== PROCESSING TYPES GRADE ERRORS GET REQUEST ===")
        
        levels = dashboards.get_levels()

        groups = list(
            Group.objects.values("groupname", "idayear__title")
            .distinct()
            .order_by("groupname")
        )
        
        courses = list(
            Group.objects.values("studycourse")
            .filter(studycourse__gt=0)
            .distinct()
            .order_by("studycourse")
        )
        
        texts = list(
            Text.objects.values("header")
            .filter(errorcheckflag=True)
            .distinct()
            .order_by("header")
        )

        text_types = list(
            TextType.objects.filter(text__errorcheckflag=True)
            .distinct()
            .order_by('idtexttype')
            .values('idtexttype', 'texttypename')
        )

        grades = list(ErrorLevel.objects.values('iderrorlevel', 'errorlevelname', 'errorlevelrussian').order_by('iderrorlevel'))

        data_count_errors = list(
            ErrorToken.objects.values(
                "iderror__iderrortag__iderrortag",
                "iderror__iderrortag__idtagparent",
                "iderror__iderrortag__tagtext",
                "iderror__iderrortag__tagtextrussian",
                "iderror__iderrortag__tagcolor",
                "iderror__iderrorlevel__iderrorlevel",
                "iderror__iderrorlevel__errorlevelname",
                "idtoken__idsentence__idtext",
            )
            .filter(
                Q(iderror__iderrortag__isnull=False) &
                Q(idtoken__idsentence__idtext__errorcheckflag=True)
            )
            .annotate(count_data=Count("iderror__iderrortag__iderrortag"))
        )

        data_on_tokens = dashboards.get_data_on_tokens(
            data_count_errors,
            id_data="iderror__iderrortag__iderrortag",
            is_unique_data=True,
            is_for_one_group=False,
        )

        data_by_grades = []
        for grade in grades:
            grade_data = [item for item in data_on_tokens if item.get('iderror__iderrorlevel__iderrorlevel') == grade['iderrorlevel']]
            processed_data = dashboards.get_data_errors(grade_data, level=0, is_sorted=True)
            data_by_grades.append(processed_data)

        print("3. data_by_grades length:", len(data_by_grades))
        for i, grade_data in enumerate(data_by_grades):
            print(f"Grade {i}: {len(grade_data)} items")

        tag_parents, dict_children = dashboards.get_dict_children()

        return render(
            request,
            "dashboard_error_types_grade.html",
            {
                "right": True,
                "levels": levels,
                "groups": groups,
                "courses": courses,
                "texts": texts,
                "text_types": text_types,
                "grades": grades,
                "data": data_by_grades,
                "tag_parents": tag_parents,
                "dict_children": dict_children,
            },
        )
    
    else:
        list_filters = json.loads(request.body)
        flag_post = list_filters['flag_post']
        
        if flag_post == 'enrollment_date':
            enrollment_date = dashboards.get_enrollment_date(list_filters)
            return JsonResponse({'enrollment_date': enrollment_date}, status=200)

        if flag_post == "choice_all":
            texts, text_types = dashboards.get_filters_for_choice_all(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_group":
            texts, text_types = dashboards.get_filters_for_choice_group(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_student":
            texts, text_types = dashboards.get_filters_for_choice_student(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_course":
            texts, text_types = dashboards.get_filters_for_choice_course(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_text":
            groups, courses, text_types = dashboards.get_filters_for_choice_text(
                list_filters
            )
            return JsonResponse(
                {"groups": groups, "courses": courses, "text_types": text_types},
                status=200,
            )

        if flag_post == "choice_text_type":
            groups, courses, texts = dashboards.get_filters_for_choice_text_type(
                list_filters
            )
            return JsonResponse(
                {"groups": groups, "courses": courses, "texts": texts}, status=200
            )
            
        if flag_post == 'update_diagrams':
            group = list_filters.get('group')
            date = list_filters.get('enrollment_date')
            surname = list_filters.get('surname')
            name = list_filters.get('name')
            patronymic = list_filters.get('patronymic')
            course = list_filters.get('course')
            text = list_filters.get('text')
            text_type = list_filters.get('text_type')
            level = int(list_filters.get('level', 0))

            print("=== FILTERS FOR TYPES GRADE DIAGRAMS ===")
            print(f"Group: {group}, Date: {date}")
            print(f"Surname: {surname}, Name: {name}, Patronymic: {patronymic}")
            print(f"Course: {course}, Text: {text}, Text_type: {text_type}")
            print(f"Level: {level}")

            academic_year_title = None
            if date:
                if " \\ " in date:
                    start_year = date.split(" \\ ")[0]
                    academic_year_title = f"{start_year}/{int(start_year) + 1}"
                else:
                    academic_year_title = date

            grades = list(ErrorLevel.objects.values('iderrorlevel', 'errorlevelname', 'errorlevelrussian').order_by('iderrorlevel'))

            data_by_grades = []

            for grade in grades:
                base_filter = Q(iderror__iderrortag__isnull=False) & \
                             Q(idtoken__idsentence__idtext__errorcheckflag=True) & \
                             Q(iderror__iderrorlevel=grade['iderrorlevel'])

                if surname and name:
                    base_filter &= Q(
                        idtoken__idsentence__idtext__idstudent__iduser__lastname__iexact=surname
                    ) & Q(
                        idtoken__idsentence__idtext__idstudent__iduser__firstname__iexact=name
                    )
                    if patronymic:
                        base_filter &= Q(
                            idtoken__idsentence__idtext__idstudent__iduser__middlename__iexact=patronymic
                        )

                if course:
                    base_filter &= Q(
                        idtoken__idsentence__idtext__idstudent__idgroup__studycourse=course
                    )

                if group and academic_year_title:
                    print(f"Applying group filter: {group}, year: {academic_year_title}")
                    base_filter &= Q(
                        idtoken__idsentence__idtext__idstudent__idgroup__groupname=group
                    ) & Q(
                        idtoken__idsentence__idtext__idstudent__idgroup__idayear__title=academic_year_title
                    )

                if text:
                    base_filter &= Q(idtoken__idsentence__idtext__header=text)

                if text_type:
                    base_filter &= Q(idtoken__idsentence__idtext__idtexttype__texttypename=text_type)

                data_count_errors = list(
                    ErrorToken.objects.filter(base_filter)
                    .values(
                        "iderror__iderrortag__iderrortag",
                        "iderror__iderrortag__idtagparent",
                        "iderror__iderrortag__tagtext",
                        "iderror__iderrortag__tagtextrussian",
                        "idtoken__idsentence__idtext",
                    )
                    .annotate(count_data=Count("iderror__iderrortag__iderrortag"))
                )

                print(f"Grade {grade['errorlevelname']}: {len(data_count_errors)} errors")

                data_on_tokens = dashboards.get_data_on_tokens(
                    data_count_errors, "iderror__iderrortag__iderrortag", True, False
                )
                grade_data = dashboards.get_data_errors(data_on_tokens, level, True)
                data_by_grades.append(grade_data)

            print(f"Final data_by_grades length: {len(data_by_grades)}")

            return JsonResponse({"data": data_by_grades}, status=200)

def chart_student_dynamics(request):		
    if request.method != 'POST':
        tags = list(ErrorTag.objects.values('iderrortag', 'tagtext', 'tagtextrussian')
                   .order_by('iderrortag').distinct())

        return render(request, 'dashboard_student_dynamics.html', 
                     {'right': True, 'tags': tags})
    
    else:
        try:
            list_filters = json.loads(request.body)
            surname = list_filters.get('surname', '')
            name = list_filters.get('name', '')
            patronymic = list_filters.get('patronymic', '')
            tag = list_filters.get('tag', '')
            checked_tag_children = list_filters.get('checked_tag_children', False)

            tags = [tag]
            if checked_tag_children:
                tags = get_tag_children(tag)

            base_text_filter = Q(errorcheckflag=True)

            if surname and name:
                student_filter = Q(idstudent__iduser__lastname__iexact=surname) & \
                               Q(idstudent__iduser__firstname__iexact=name)
                if patronymic:
                    student_filter &= Q(idstudent__iduser__middlename__iexact=patronymic)
                base_text_filter &= student_filter

            error_query = Error.objects.filter(
                iderrortag__in=tags,
                errortoken__idtoken__idsentence__idtext__in=Text.objects.filter(base_text_filter)
            )

            data_count_errors = list(
                error_query.values(
                    'errortoken__idtoken__idsentence__idtext__createdate'
                ).annotate(
                    count_data=Count('errortoken__idtoken__idsentence__idtext__createdate'),
                    text_id=F('errortoken__idtoken__idsentence__idtext__idtext')
                ).order_by('errortoken__idtoken__idsentence__idtext__createdate')
            )

            text_ids_with_errors = set(
                error_query.values_list(
                    'errortoken__idtoken__idsentence__idtext__idtext', 
                    flat=True
                ).distinct()
            )

            texts_without_errors = Text.objects.filter(
                base_text_filter
            ).exclude(
                idtext__in=text_ids_with_errors
            ).values(
                'createdate', 'idtext'
            ).annotate(
                count_data=Value(0, output_field=IntegerField())
            )

            combined_data = list(data_count_errors) + list(texts_without_errors)

            combined_data = [item for item in combined_data if item.get('errortoken__idtoken__idsentence__idtext__createdate') or item.get('createdate')]
            combined_data = sorted(combined_data, 
                                 key=lambda x: x.get('errortoken__idtoken__idsentence__idtext__createdate') or x.get('createdate'))
            
            return JsonResponse({
                'data': combined_data
            }, status=200)
            
        except Exception as e:
            print(f"Error in chart_student_dynamics: {str(e)}")
            return JsonResponse({
                'error': str(e),
                'data': []
            }, status=500)

# Функция для получения дочерних тегов
def get_tag_children(tag_id):
    return list(ErrorTag.objects.filter(
        idtagparent=tag_id
    ).values_list('iderrortag', flat=True))


def chart_groups_errors(request):		
    if request.method != 'POST':
        tags = list(ErrorTag.objects.values('iderrortag', 'tagtext', 'tagtextrussian')
                   .order_by('iderrortag').distinct())
        groups = list(Group.objects.select_related('idayear').values(
            'idgroup', 'groupname', 'idayear__title'
        ).distinct().order_by('-idayear__title', 'groupname'))
        
        return render(request, 'dashboard_error_groups.html', {
            'right': True, 
            'tags': tags,
            'groups': groups
        })
    else:
        try:
            list_filters = json.loads(request.body)
            text = list_filters.get('text', '')
            text_type = list_filters.get('text_type', '')
            groups = list_filters.get('group', [])
            tag = list_filters.get('tag', '')
            checked_tag_children = list_filters.get('checked_tag_children', False)

            tags_list = [tag]
            if checked_tag_children:
                tags_list = get_tag_children(tag)

            base_texts_query = Text.objects.filter(errorcheckflag=True)

            if groups:
                base_texts_query = base_texts_query.filter(idstudent__idgroup__in=groups)

            texts_with_errors = base_texts_query.filter(
                idtext__in=Error.objects.filter(
                    iderrortag__in=tags_list,
                    errortoken__idtoken__idsentence__idtext__in=base_texts_query
                ).values_list('errortoken__idtoken__idsentence__idtext', flat=True).distinct()
            )

            errors_data = list(
                Error.objects.filter(
                    iderrortag__in=tags_list,
                    errortoken__idtoken__idsentence__idtext__in=texts_with_errors
                )
                .values(
                    groupname=F('errortoken__idtoken__idsentence__idtext__idstudent__idgroup__groupname'),
                    idayear__title=F('errortoken__idtoken__idsentence__idtext__idstudent__idgroup__idayear__title')
                )
                .annotate(
                    count_data=Count('iderror')
                )
                .order_by('-count_data')
            )

            if groups:
                groups_with_errors = set([item['groupname'] for item in errors_data])
                groups_without_errors = Group.objects.filter(
                    idgroup__in=groups
                ).exclude(
                    groupname__in=groups_with_errors
                ).values('groupname', 'idayear__title')
                
                for group in groups_without_errors:
                    errors_data.append({
                        'groupname': group['groupname'],
                        'idayear__title': group['idayear__title'],
                        'count_data': 0
                    })

            texts_query = texts_with_errors
            if text_type:
                texts_query = texts_query.filter(idtexttype=text_type)
            if text:
                texts_query = texts_query.filter(header=text)
            
            texts = list(
                texts_query.values('idtext', 'header').distinct().order_by('header')
            )

            text_types_query = TextType.objects.filter(
                text__in=texts_with_errors
            )
            if text:
                text_types_query = text_types_query.filter(text__header=text)
            if text_type:
                text_types_query = text_types_query.filter(idtexttype=text_type)
            
            text_types = list(
                text_types_query.values('idtexttype', 'texttypename').distinct().order_by('texttypename')
            )
            
            return JsonResponse({
                'data': errors_data, 
                'texts': texts, 
                'text_types': text_types
            }, status=200)
            
        except Exception as e:
            print(f"Error in chart_groups_errors: {str(e)}")
            import traceback
            print(traceback.format_exc())
            return JsonResponse({
                'error': str(e),
                'data': [],
                'texts': [],
                'text_types': []
            }, status=500)

def chart_emotions_errors(request):		
    if request.method != 'POST':
        print("=== PROCESSING EMOTIONS GET REQUEST ===")

        levels = dashboards.get_levels()

        emotions = list(Emotion.objects.values('idemotion', 'emotionname'))

        groups = list(
            Group.objects.values("groupname", "idayear__title")
            .distinct()
            .order_by("groupname")
        )

        courses = list(
            Group.objects.values("studycourse")
            .filter(studycourse__gt=0)
            .distinct()
            .order_by("studycourse")
        )

        texts = list(
            Text.objects.values("header")
            .filter(errorcheckflag=True)
            .distinct()
            .order_by("header")
        )

        text_types = list(
            TextType.objects.filter(text__errorcheckflag=True)
            .distinct()
            .order_by('idtexttype')
            .values('idtexttype', 'texttypename')
        )

        data_count_errors = list(
            ErrorToken.objects.values(
                "iderror__iderrortag__iderrortag",
                "iderror__iderrortag__idtagparent",
                "iderror__iderrortag__tagtext",
                "iderror__iderrortag__tagtextrussian",
                "iderror__iderrortag__tagcolor",
                "idtoken__idsentence__idtext",
            )
            .filter(
                Q(iderror__iderrortag__isnull=False) &
                Q(idtoken__idsentence__idtext__errorcheckflag=True) &
                Q(idtoken__idsentence__idtext__idemotion__isnull=False)  
            )
            .annotate(count_data=Count("iderror__iderrortag__iderrortag"))
        )

        data_on_tokens = dashboards.get_data_on_tokens(
            data_count_errors,
            id_data="iderror__iderrortag__iderrortag",
            is_unique_data=True,
            is_for_one_group=False,
        )


        data = dashboards.get_data_errors(data_on_tokens, level=0, is_sorted=True)

        tag_parents, dict_children = dashboards.get_dict_children()

        return render(
            request,
            "dashboard_error_emotions.html",
            {
                "right": True,
                "levels": levels,
                "emotions": emotions,
                "groups": groups,
                "courses": courses,
                "texts": texts,
                "text_types": text_types,
                "data": data,
                "tag_parents": tag_parents,
                "dict_children": dict_children,
            },
        )
    
    else:
        list_filters = json.loads(request.body)
        flag_post = list_filters['flag_post']
        
        if flag_post == 'enrollment_date':
            enrollment_date = dashboards.get_enrollment_date(list_filters)
            return JsonResponse({'enrollment_date': enrollment_date}, status=200)

        if flag_post == "choice_all":
            texts, text_types = dashboards.get_filters_for_choice_all(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_group":
            texts, text_types = dashboards.get_filters_for_choice_group(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_student":
            texts, text_types = dashboards.get_filters_for_choice_student(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_course":
            texts, text_types = dashboards.get_filters_for_choice_course(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_text":
            groups, courses, text_types = dashboards.get_filters_for_choice_text(
                list_filters
            )
            return JsonResponse(
                {"groups": groups, "courses": courses, "text_types": text_types},
                status=200,
            )

        if flag_post == "choice_text_type":
            groups, courses, texts = dashboards.get_filters_for_choice_text_type(
                list_filters
            )
            return JsonResponse(
                {"groups": groups, "courses": courses, "texts": texts}, status=200
            )
            
        if flag_post == 'update_diagrams':
            group = list_filters.get('group')
            date = list_filters.get('enrollment_date')
            surname = list_filters.get('surname')
            name = list_filters.get('name')
            patronymic = list_filters.get('patronymic')
            course = list_filters.get('course')
            text = list_filters.get('text')
            text_type = list_filters.get('text_type')
            emotion = list_filters.get('emotion')
            level = int(list_filters.get('level', 0))

            print("=== FILTERS FOR EMOTIONS DIAGRAMS ===")
            print(f"Emotion: {emotion}")
            print(f"Group: {group}, Date: {date}")
            print(f"Surname: {surname}, Name: {name}, Patronymic: {patronymic}")
            print(f"Course: {course}, Text: {text}, Text_type: {text_type}")
            print(f"Level: {level}")

            academic_year_title = None
            if date:
                if " \\ " in date:
                    start_year = date.split(" \\ ")[0]
                    academic_year_title = f"{start_year}/{int(start_year) + 1}"
                else:
                    academic_year_title = date

            base_filter = Q(iderror__iderrortag__isnull=False) & \
                         Q(idtoken__idsentence__idtext__errorcheckflag=True) & \
                         Q(idtoken__idsentence__idtext__idemotion=emotion)

            if surname and name:
                base_filter &= Q(
                    idtoken__idsentence__idtext__idstudent__iduser__lastname__iexact=surname
                ) & Q(
                    idtoken__idsentence__idtext__idstudent__iduser__firstname__iexact=name
                )
                if patronymic:
                    base_filter &= Q(
                        idtoken__idsentence__idtext__idstudent__iduser__middlename__iexact=patronymic
                    )

            if course:
                base_filter &= Q(
                    idtoken__idsentence__idtext__idstudent__idgroup__studycourse=course
                )

            if group and academic_year_title:
                print(f"Applying group filter: {group}, year: {academic_year_title}")
                base_filter &= Q(
                    idtoken__idsentence__idtext__idstudent__idgroup__groupname=group
                ) & Q(
                    idtoken__idsentence__idtext__idstudent__idgroup__idayear__title=academic_year_title
                )

            if text:
                base_filter &= Q(idtoken__idsentence__idtext__header=text)

            if text_type:
                base_filter &= Q(idtoken__idsentence__idtext__idtexttype__texttypename=text_type)

            data_count_errors = list(
                ErrorToken.objects.filter(base_filter)
                .values(
                    "iderror__iderrortag__iderrortag",
                    "iderror__iderrortag__idtagparent",
                    "iderror__iderrortag__tagtext",
                    "iderror__iderrortag__tagtextrussian",
                    "idtoken__idsentence__idtext",
                )
                .annotate(count_data=Count("iderror__iderrortag__iderrortag"))
            )

            print(f"Filtered data_count_errors length: {len(data_count_errors)}")

            data_on_tokens = dashboards.get_data_on_tokens(
                data_count_errors, "iderror__iderrortag__iderrortag", True, False
            )
            data = dashboards.get_data_errors(data_on_tokens, level, True)

            print(f"Final data length: {len(data)}")

            return JsonResponse({"data_type_errors": data}, status=200)

def chart_self_rating_errors(request):		
    if request.method != 'POST':

        levels = dashboards.get_levels()

        self_ratings = []
        for value, label in Text.TASK_RATES:
            texts_with_rating = Text.objects.filter(
                selfrating=value, 
                errorcheckflag=True
            ).exists()
            if texts_with_rating:
                self_ratings.append({
                    'selfrating': value,
                    'selfrating_text': label
                })

        if not self_ratings:
            self_ratings = [{'selfrating': value, 'selfrating_text': label} 
                           for value, label in Text.TASK_RATES]

        groups = list(
            Group.objects.values("groupname", "idayear__title")
            .distinct()
            .order_by("groupname")
        )

        courses = list(
            Group.objects.values("studycourse")
            .filter(studycourse__gt=0)
            .distinct()
            .order_by("studycourse")
        )

        texts = list(
            Text.objects.values("header")
            .filter(errorcheckflag=True)
            .distinct()
            .order_by("header")
        )

        text_types = list(
            TextType.objects.filter(text__errorcheckflag=True)
            .distinct()
            .order_by('idtexttype')
            .values('idtexttype', 'texttypename')
        )

        data_count_errors = list(
            ErrorToken.objects.values(
                "iderror__iderrortag__iderrortag",
                "iderror__iderrortag__idtagparent",
                "iderror__iderrortag__tagtext",
                "iderror__iderrortag__tagtextrussian",
                "iderror__iderrortag__tagcolor",
                "idtoken__idsentence__idtext",
            )
            .filter(
                Q(iderror__iderrortag__isnull=False) &
                Q(idtoken__idsentence__idtext__errorcheckflag=True) &
                Q(idtoken__idsentence__idtext__selfrating__isnull=False)  
            )
            .annotate(count_data=Count("iderror__iderrortag__iderrortag"))
        )

        data_on_tokens = dashboards.get_data_on_tokens(
            data_count_errors,
            id_data="iderror__iderrortag__iderrortag",
            is_unique_data=True,
            is_for_one_group=False,
        )

        data = dashboards.get_data_errors(data_on_tokens, level=0, is_sorted=True)

        tag_parents, dict_children = dashboards.get_dict_children()


        return render(
            request,
            "dashboard_error_self_rating.html",
            {
                "right": True,
                "levels": levels,
                "self_ratings": self_ratings,
                "groups": groups,
                "courses": courses,
                "texts": texts,
                "text_types": text_types,
                "data": data,
                "tag_parents": tag_parents,
                "dict_children": dict_children,
            },
        )
    
    else:
        list_filters = json.loads(request.body)
        flag_post = list_filters['flag_post']
        
        if flag_post == 'enrollment_date':
            enrollment_date = dashboards.get_enrollment_date(list_filters)
            return JsonResponse({'enrollment_date': enrollment_date}, status=200)

        if flag_post == "choice_all":
            texts, text_types = dashboards.get_filters_for_choice_all(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_group":
            texts, text_types = dashboards.get_filters_for_choice_group(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_student":
            texts, text_types = dashboards.get_filters_for_choice_student(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_course":
            texts, text_types = dashboards.get_filters_for_choice_course(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_text":
            groups, courses, text_types = dashboards.get_filters_for_choice_text(
                list_filters
            )
            return JsonResponse(
                {"groups": groups, "courses": courses, "text_types": text_types},
                status=200,
            )

        if flag_post == "choice_text_type":
            groups, courses, texts = dashboards.get_filters_for_choice_text_type(
                list_filters
            )
            return JsonResponse(
                {"groups": groups, "courses": courses, "texts": texts}, status=200
            )
            
        if flag_post == 'update_diagrams':
            group = list_filters.get('group')
            date = list_filters.get('enrollment_date')
            surname = list_filters.get('surname')
            name = list_filters.get('name')
            patronymic = list_filters.get('patronymic')
            course = list_filters.get('course')
            text = list_filters.get('text')
            text_type = list_filters.get('text_type')
            self_rating = list_filters.get('self_rating')
            level = int(list_filters.get('level', 0))

            academic_year_title = None
            if date:
                if " \\ " in date:
                    start_year = date.split(" \\ ")[0]
                    academic_year_title = f"{start_year}/{int(start_year) + 1}"
                else:
                    academic_year_title = date

            base_filter = Q(iderror__iderrortag__isnull=False) & \
                         Q(idtoken__idsentence__idtext__errorcheckflag=True) & \
                         Q(idtoken__idsentence__idtext__selfrating=self_rating)  

            if surname and name:
                base_filter &= Q(
                    idtoken__idsentence__idtext__idstudent__iduser__lastname__iexact=surname
                ) & Q(
                    idtoken__idsentence__idtext__idstudent__iduser__firstname__iexact=name
                )
                if patronymic:
                    base_filter &= Q(
                        idtoken__idsentence__idtext__idstudent__iduser__middlename__iexact=patronymic
                    )

            if course:
                base_filter &= Q(
                    idtoken__idsentence__idtext__idstudent__idgroup__studycourse=course
                )

            if group and academic_year_title:
                print(f"Applying group filter: {group}, year: {academic_year_title}")
                base_filter &= Q(
                    idtoken__idsentence__idtext__idstudent__idgroup__groupname=group
                ) & Q(
                    idtoken__idsentence__idtext__idstudent__idgroup__idayear__title=academic_year_title
                )

            if text:
                base_filter &= Q(idtoken__idsentence__idtext__header=text)

            if text_type:
                base_filter &= Q(idtoken__idsentence__idtext__idtexttype__texttypename=text_type)

            data_count_errors = list(
                ErrorToken.objects.filter(base_filter)
                .values(
                    "iderror__iderrortag__iderrortag",
                    "iderror__iderrortag__idtagparent",
                    "iderror__iderrortag__tagtext",
                    "iderror__iderrortag__tagtextrussian",
                    "idtoken__idsentence__idtext",
                )
                .annotate(count_data=Count("iderror__iderrortag__iderrortag"))
            )

            print(f"Filtered data_count_errors length: {len(data_count_errors)}")

            data_on_tokens = dashboards.get_data_on_tokens(
                data_count_errors, "iderror__iderrortag__iderrortag", True, False
            )
            data = dashboards.get_data_errors(data_on_tokens, level, True)

            print(f"Final data length: {len(data)}")

            return JsonResponse({"data_type_errors": data}, status=200)
        
def chart_relation_assessment_self_rating(request):    
    if request.method != 'POST':
        return render(request, 'dashboard_assessment_self_rating.html', {'right': True})
    
    else:
        try:
            list_filters = json.loads(request.body)
            surname = list_filters.get('surname', '').strip()
            name = list_filters.get('name', '').strip()
            patronymic = list_filters.get('patronymic', '').strip()
            text_type = list_filters.get('text_type', '').strip()

            base_filter = Q(selfrating__gt=0) & Q(textgrade__gt=0) & Q(errorcheckflag=True)

            if surname and name:
                base_filter &= Q(idstudent__iduser__lastname__iexact=surname) & Q(idstudent__iduser__firstname__iexact=name)
                if patronymic:
                    base_filter &= Q(idstudent__iduser__middlename__iexact=patronymic)

            if text_type:
                base_filter &= Q(idtexttype__texttypename=text_type)

            texts_query = Text.objects.filter(base_filter)

            data_relation = list(
                texts_query.values('textgrade', 'selfrating').distinct()
            )

            assessment_types = Text.TASK_RATES
            
            for data in data_relation:
                self_rating_idx = data["selfrating"]
                text_grade_idx = data["textgrade"]

                data["self_rating_text"] = (
                    assessment_types[self_rating_idx - 1][1] 
                    if 0 < self_rating_idx <= len(assessment_types) 
                    else str(self_rating_idx)
                )
                
                data["assessment_text"] = (
                    assessment_types[text_grade_idx - 1][1] 
                    if 0 < text_grade_idx <= len(assessment_types) 
                    else str(text_grade_idx)
                )

            if surname and name:
                student_filter = Q(text__idstudent__iduser__lastname__iexact=surname) & Q(text__idstudent__iduser__firstname__iexact=name)
                if patronymic:
                    student_filter &= Q(text__idstudent__iduser__middlename__iexact=patronymic)
                
                text_types = list(
                    TextType.objects.filter(
                        student_filter &
                        Q(text__selfrating__gt=0) & 
                        Q(text__textgrade__gt=0) & 
                        Q(text__errorcheckflag=True)
                    )
                    .values('idtexttype', 'texttypename')
                    .distinct()
                    .order_by('texttypename')
                )
            else:
                text_types = list(
                    TextType.objects.filter(
                        Q(text__selfrating__gt=0) & 
                        Q(text__textgrade__gt=0) & 
                        Q(text__errorcheckflag=True)
                    )
                    .values('idtexttype', 'texttypename')
                    .distinct()
                    .order_by('texttypename')
                )
                
            return JsonResponse({
                'data': data_relation, 
                'text_types': text_types
            }, status=200)
            
        except Exception as e:
            return JsonResponse({
                'error': str(e),
                'data': [],
                'text_types': []
            }, status=500)








# ____________________________________________________________________________






# ПОИСК ЗАВИСИМОСТЕЙ

def relation_emotions_self_rating(request):	
	if request.method != 'POST':
		languages = ['Deustache']
		courses = list(
			Group.objects.values('studycourse').filter(studycourse__gt=0).distinct().order_by(
				'studycourse'))
		groups = list(Group.objects.values('groupname').distinct().order_by('groupname'))
		
		data_relation = list(
			Text.objects.values('idemotion', 'selfrating').filter(
				Q(idemotion__isnull=False) & Q(selfrating__gt=0) & ~Q(idemotion=2)))
		
		data, relation, data_fisher = dashboards.get_stat(data_relation, 'idemotion', 'idemotion__emotionname',
								  'selfrating', 'self_rating_text', True)
		
		return render(request, 'relation_emotions_self_rating.html', {'right': True, 'languages': languages,
									      'courses': courses, 'groups': groups,
									      'data_relation': data, 'relation': relation,
									      'data_fisher': data_fisher})
	else:
		list_filters = json.loads(request.body)
		flag_post = list_filters['flag_post']
		
		if flag_post == 'enrollment_date':
			group = list_filters['group']
			enrollment_date = list(
				Group.objects.values('enrollment_date').filter(groupname=group).distinct().order_by(
					'enrollment_date'))
			
			for date in enrollment_date:
				date['enrollment_date'] = str(date['enrollment_date'].year) + ' \ ' \
								+ str(date['enrollment_date'].year + 1)
				
			return JsonResponse({'enrollment_date': enrollment_date}, status=200)
			
		if flag_post == 'course':
			course = list_filters['course']
			
			data_relation = list(Text.objects.values('idemotion', 'selfrating').filter(
				Q(emotional__isnull=False) & Q(self_rating__gt=0) & Q(tbltextgroup__group__course_number=course) & ~Q(
					emotional=2)))
			
		if flag_post == 'group':
			group = list_filters['group']
			date = list_filters['date']
			group_date = date[:4] + '-09-01'
			
			data_relation = list(
				Text.objects.values('idemotion', 'selfrating').filter(
					Q(emotional__isnull=False) & Q(self_rating__gt=0) & Q(
						tbltextgroup__group__group_name=group) & Q(
						tbltextgroup__group__enrollment_date=group_date) & ~Q(emotional=2)))
			
		data, relation, data_fisher = dashboards.get_stat(data_relation, 'idemotion', 'idemotion__emotionname',
								  'selfrating', 'self_rating_text', True)
		
		return JsonResponse({'data_relation': data, 'relation': relation, 'data_fisher': data_fisher}, status=200)


def relation_emotions_assessment(request):	
	if request.method != 'POST':
		languages = ['Deustache']
		courses = list(
			Group.objects.values('studycourse').filter(studycourse__gt=0).distinct().order_by(
				'studycourse'))
		groups = list(Group.objects.values('groupname').distinct().order_by('groupname'))
		
		data_relation = list(
			Text.objects.values('idemotion', 'textgrade').filter(
				Q(idemotion__isnull=False) & Q(textgrade__gt=0) & ~Q(idemotion=2)))
		
		data, relation, data_fisher = dashboards.get_stat(data_relation, 'idemotion', 'idemotion__emotionname',
								  'textgrade', 'assessment_text', True)
		
		return render(request, 'relation_emotions_assessment.html', {'right': True, 'languages': languages,
									     'courses': courses, 'groups': groups,
									     'data_relation': data, 'relation': relation,
									     'data_fisher': data_fisher})
	else:
		list_filters = json.loads(request.body)
		flag_post = list_filters['flag_post']
		
		if flag_post == 'enrollment_date':
			group = list_filters['group']
			enrollment_date = list(
				Group.objects.values('enrollment_date').filter(groupname=group).distinct().order_by(
					'enrollment_date'))
			
			for date in enrollment_date:
				date['enrollment_date'] = str(date['enrollment_date'].year) + ' \ ' \
								+ str(date['enrollment_date'].year + 1)
				
			return JsonResponse({'enrollment_date': enrollment_date}, status=200)
			
		if flag_post == 'course':
			course = list_filters['course']
			
			data_relation = list(
				Text.objects.values('idemotion', 'textgrade').filter(
					Q(emotional__isnull=False) & Q(assessment__gt=0) & Q(
						tbltextgroup__group__course_number=course) & ~Q(emotional=2)))
			
		if flag_post == 'group':
			group = list_filters['group']
			date = list_filters['date']
			group_date = date[:4] + '-09-01'
			
			data_relation = list(Text.objects.values('idemotion', 'textgrade').filter(
				Q(emotional__isnull=False) & Q(assessment__gt=0) & Q(tbltextgroup__group__group_name=group) & Q(
					tbltextgroup__group__enrollment_date=group_date) & ~Q(emotional=2)))
			
		data, relation, data_fisher = dashboards.get_stat(data_relation, 'idemotion', 'idemotion__emotionname',
								  'textgrade', 'assessment_text', True)
		
		return JsonResponse({'data_relation': data, 'relation': relation, 'data_fisher': data_fisher}, status=200)


def relation_self_rating_assessment(request):	
	if request.method != 'POST':
		languages = ['Deustache']
		courses = list(
			Group.objects.values('studycourse').filter(studycourse__gt=0).distinct().order_by(
				'studycourse'))
		groups = list(Group.objects.values('groupname').distinct().order_by('groupname'))
		
		data_relation = list(Text.objects.values('selfrating', 'textgrade').filter(
			Q(selfrating__gt=0) & Q(textgrade__gt=0)))
		
		data, relation, data_fisher = dashboards.get_stat(data_relation, 'selfrating', 'self_rating_text',
								  'textgrade', 'assessment_text', False)
		
		return render(request, 'relation_self_rating_assessment.html', {'right': True, 'languages': languages,
										'courses': courses, 'groups': groups,
										'data_relation': data, 'relation': relation,
										'data_fisher': data_fisher})
	else:
		list_filters = json.loads(request.body)
		flag_post = list_filters['flag_post']
		
		if flag_post == 'enrollment_date':
			group = list_filters['group']
			enrollment_date = list(
				Group.objects.values('enrollment_date').filter(groupname=group).distinct().order_by(
					'enrollment_date'))
			
			for date in enrollment_date:
				date['enrollment_date'] = str(date['enrollment_date'].year) + ' \ ' \
								+ str(date['enrollment_date'].year + 1)
				
			return JsonResponse({'enrollment_date': enrollment_date}, status=200)
			
		if flag_post == 'course':
			course = list_filters['course']
			
			data_relation = list(Text.objects.values('selfrating', 'textgrade').filter(
				Q(self_rating__gt=0) & Q(assessment__gt=0) & Q(tbltextgroup__group__course_number=course)))
			
		if flag_post == 'group':
			group = list_filters['group']
			date = list_filters['date']
			group_date = date[:4] + '-09-01'
			
			data_relation = list(Text.objects.values('selfrating', 'textgrade').filter(
				Q(self_rating__gt=0) & Q(assessment__gt=0) & Q(tbltextgroup__group__group_name=group) & Q(
					tbltextgroup__group__enrollment_date=group_date)))
			
		data, relation, data_fisher = dashboards.get_stat(data_relation, 'selfrating', 'self_rating_text',
								  'textgrade', 'assessment_text', False)
		
		return JsonResponse({'data_relation': data, 'relation': relation, 'data_fisher': data_fisher}, status=200)


def relation_course_errors(request):	
	if request.method != 'POST':
		languages = ['Deustache']
		tags = list(Error.objects.values('iderrortag__iderrortag', 'iderrortag__tagtext', 'iderrortag__tagtextrussian').order_by('iderrortag__iderrortag'))
		
		return render(request, 'relation_course_errors.html', {'right': True, 'languages': languages, 'tags': tags})
	else:
		list_filters = json.loads(request.body)
		flag_post = list_filters['flag_post']
		tag = list_filters['tag']
		checked_tag_children = list_filters['checked_tag_children']
		
		tags = [tag]
		if checked_tag_children:
			tags = dashboards.get_tag_children(tag)
			
		if flag_post == 'courses':
			data_relation = list(
				Error.objects.values('sentence__text_id__tbltextgroup__group__course_number').filter(
					Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(tag__id_tag__in=tags) & Q(
						sentence__text_id__tbltextgroup__group__course_number__isnull=False)).annotate(
					count_data=Count('sentence__text_id__tbltextgroup__group__course_number')))
			
		if flag_post == 'students':
			data_relation = list(Error.objects.values('sentence__text_id__user',
								      'sentence__text_id__tbltextgroup__group__course_number').filter(
				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(tag__id_tag__in=tags) & Q(
					sentence__text_id__tbltextgroup__group__course_number__isnull=False)).annotate(
				count_data=Count('sentence__text_id__tbltextgroup__group__course_number')))
			
		if flag_post == 'groups':
			data_relation = list(Error.objects.values('sentence__text_id__tbltextgroup__group',
								      'sentence__text_id__tbltextgroup__group__course_number').filter(
				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
					sentence__text_id__tbltextgroup__group__isnull=False) & Q(tag__id_tag__in=tags)).annotate(
				count_data=Count('sentence__text_id__tbltextgroup__group__course_number')))
			
		course = []
		count_errors = []
		
		for data in data_relation:
			course.append(data['sentence__text_id__tbltextgroup__group__course_number'])
			count_errors.append(data['count_data'])
			
		critical_stat_level = 0.05
		n = len(course)
		
		if n > 1:
			if len(set(course)) == 1 or len(set(count_errors)) == 1:
				relation = {'result': 'один из параметров константа', 'stat': 'None', 'pvalue': 'None', 'N': n}
				
			else:
				result = scipy.stats.spearmanr(course, count_errors)
				
				t = abs(result.statistic) * np.sqrt((n-2) / (1 - result.statistic * result.statistic))
				t_critical = scipy.stats.t.ppf(1-critical_stat_level/2, n-2)
				
				if t < t_critical:
					worth = 'корреляция статистически не значимая'
				else:
					worth = 'статистически значимая корреляция'
					
				if np.isnan(result.pvalue):
					pvalue = 'Nan'
				else:
					pvalue = result.pvalue
					
				if result.statistic == 0:
					relation = {'result': f'связь отсутствует  ({worth})', 'stat': result.statistic, 'pvalue': pvalue,
						    'N': n}
				elif result.statistic >= 0.75:
					relation = {'result': f'очень высокая положительная связь ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
				elif 0.5 <= result.statistic < 0.75:
					relation = {'result': f'высокая положительная связь  ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
				elif 0.25 <= result.statistic < 0.5:
					relation = {'result': f'средняя положительная связь  ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
				elif 0 < result.statistic < 0.25:
					relation = {'result': f'слабая положительная связь  ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
				elif -0.25 <= result.statistic < 0:
					relation = {'result': f'слабая отрицательная связь  ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
				elif -0.5 <= result.statistic < -0.25:
					relation = {'result': f'средняя отрицательная связь  ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
				elif -0.75 <= result.statistic < -0.5:
					relation = {'result': f'высокая отрицательная связь  ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
				else:
					relation = {'result': f'очень высокая отрицательная связь  ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
		else:
			relation = {'result': '-', 'stat': '-', 'pvalue': '-', 'N': n}
			
		return JsonResponse({'data_relation': data_relation, 'relation': relation}, status=200)